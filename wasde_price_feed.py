"""
wasde_price_feed.py
===================
World Bank Pink Sheet + IMF commodity price ingestor.
No API key required. Provides 30-day price movement data for WASDE backtest.

Commodities tracked: Wheat, Corn, Soybeans, Cotton, Rice
Sources:
  - World Bank Pink Sheet (monthly Excel, stable URL pattern)
  - IMF Primary Commodity Prices (JSON API, no key)

Outputs:
  data/wasde/price-history.json  — rolling 24-month price data
  data/wasde/price-latest.json   — most recent month per commodity

WASDE backtest use: compare price T+30 vs price T for each release month.
"""
import json, sys, io, urllib.request
from datetime import datetime, timedelta
from pathlib import Path

sys.path.insert(0, "G:/My Drive/Projects/_studio")
OUT_DIR = Path("G:/My Drive/Projects/_studio/data/wasde")
OUT_DIR.mkdir(parents=True, exist_ok=True)

# ── COMMODITY MAPPINGS ────────────────────────────────────────────
# IMF Primary Commodity Price API series codes
IMF_CODES = {
    "Wheat":    "PWHEAMT",   # Wheat, US HRW
    "Corn":     "PMAIZE",    # Maize (corn), US No.2
    "Soybeans": "PSOYB",     # Soybeans, US
    "Cotton":   "PCOTTIND",  # Cotton, Cotton Outlook 'A Index'
    "Rice":     "PRICENPQ",  # Rice, Thai 5% broken
}

# World Bank Pink Sheet Excel column names (approximate — verified per sheet)
WB_COLUMNS = {
    "Wheat":    "Wheat, US HRW",
    "Corn":     "Maize",
    "Soybeans": "Soybeans",
    "Cotton":   "Cotton, A Index",
    "Rice":     "Rice, Thai 5%",
}

# ── IMF PRICE FETCHER ─────────────────────────────────────────────
def fetch_imf_prices(months_back: int = 24) -> dict:
    """
    Fetch commodity prices from IMF Primary Commodity Prices API.
    Free, no key, JSON. Returns {commodity: [{date, price_usd}, ...]}
    """
    results = {}
    base = "https://www.imf.org/external/datamapper/api/v1"

    for commodity, code in IMF_CODES.items():
        url = f"{base}/{code}/PWORLD"
        try:
            req = urllib.request.Request(url, headers={"Accept": "application/json",
                                                        "User-Agent": "StudioAgent/1.0"})
            with urllib.request.urlopen(req, timeout=15) as r:
                data = json.loads(r.read())
            # IMF response: data.values.PWORLD = {year: {period: value}}
            values = (data.get("values", {})
                         .get(code, {})
                         .get("PWORLD", {}))
            # Flatten to list of {date, price}
            prices = []
            for year, periods in sorted(values.items()):
                for period, val in sorted(periods.items()):
                    if val is not None:
                        prices.append({
                            "date":      f"{year}-{period.zfill(2)}",
                            "price_usd": round(float(val), 4),
                            "source":    "IMF"
                        })
            # Keep last N months
            results[commodity] = prices[-months_back:] if prices else []
            print(f"  IMF {commodity:12}: {len(results[commodity])} months fetched")
        except Exception as e:
            print(f"  IMF {commodity:12}: FAIL — {str(e)[:60]}")
            results[commodity] = []

    return results

# ── WORLD BANK PINK SHEET FETCHER ────────────────────────────────
def fetch_wb_pink_sheet() -> dict:
    """
    Download World Bank Pink Sheet Excel and extract commodity prices.
    URL pattern: stable document ID, monthly file updated each month.
    Falls back gracefully if openpyxl not available.
    """
    # Stable World Bank CMO Pink Sheet URL
    WB_URL = ("https://thedocs.worldbank.org/en/doc/"
              "18675f1d1639c7a34d463f59263ba0a2-0050012025/"
              "related/CMO-Historical-Data-Monthly.xlsx")
    results = {}
    try:
        import openpyxl
    except ImportError:
        print("  WB Pink Sheet: openpyxl not installed — run: pip install openpyxl")
        return results

    try:
        req = urllib.request.Request(WB_URL, headers={"User-Agent": "StudioAgent/1.0"})
        with urllib.request.urlopen(req, timeout=30) as r:
            wb_bytes = r.read()
        wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), read_only=True, data_only=True)

        # Pink Sheet monthly data is on sheet "Monthly Prices"
        sheet_name = next((s for s in wb.sheetnames
                           if "monthly" in s.lower() or "prices" in s.lower()), wb.sheetnames[0])
        ws = wb[sheet_name]

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            print("  WB Pink Sheet: empty sheet")
            return results

        # Find header row (contains commodity names)
        header_row = None
        header_idx = 0
        for i, row in enumerate(rows[:10]):
            if row and any(isinstance(c, str) and "wheat" in str(c).lower() for c in row):
                header_row = row
                header_idx = i
                break

        if header_row is None:
            print("  WB Pink Sheet: header not found")
            return results

        # Map commodity to column index
        col_map = {}
        for commodity, col_name in WB_COLUMNS.items():
            for j, cell in enumerate(header_row):
                if cell and col_name.lower() in str(cell).lower():
                    col_map[commodity] = j
                    break

        print(f"  WB Pink Sheet: found columns for {list(col_map.keys())}")

        # Extract date + price rows
        date_col = 0  # First column is date
        for commodity, col_idx in col_map.items():
            prices = []
            for row in rows[header_idx + 1:]:
                if not row or not row[date_col]:
                    continue
                date_val = row[date_col]
                price_val = row[col_idx] if len(row) > col_idx else None
                if price_val and isinstance(price_val, (int, float)):
                    # Normalize date: "2025M03" -> "2025-03"
                    raw_date = str(date_val)[:7].replace("M", "-")
                    prices.append({
                        "date":      raw_date,
                        "price_usd": round(float(price_val), 4),
                        "source":    "WorldBank"
                    })
            results[commodity] = prices[-24:]
            print(f"  WB {commodity:12}: {len(results[commodity])} months")

    except Exception as e:
        print(f"  WB Pink Sheet: FAIL — {str(e)[:80]}")

    return results

# ── BACKTEST PRICE LOOKUP ─────────────────────────────────────────
def get_price_movement(commodity: str, release_date: str,
                        price_history: dict, days: int = 30) -> dict:
    """
    For a WASDE release date, find price at release and price 30 days later.
    Returns direction: BULLISH (price rose) / BEARISH (price fell) / NEUTRAL.
    """
    prices = price_history.get(commodity, [])
    if not prices:
        return {"direction": "UNKNOWN", "reason": "no price data"}

    # Find release month price
    release_ym = release_date[:7]  # YYYY-MM
    after_date = (datetime.strptime(release_date, "%Y-%m-%d")
                  + timedelta(days=days)).strftime("%Y-%m")

    release_price = next((p["price_usd"] for p in prices
                          if p["date"][:7] == release_ym), None)
    after_price   = next((p["price_usd"] for p in prices
                          if p["date"][:7] == after_date), None)

    if not release_price or not after_price:
        return {"direction": "UNKNOWN",
                "reason": f"missing price for {release_ym} or {after_date}"}

    pct_change = (after_price - release_price) / release_price * 100
    direction = "BULLISH" if pct_change > 1.0 else \
                "BEARISH" if pct_change < -1.0 else "NEUTRAL"

    return {
        "direction":      direction,
        "release_price":  release_price,
        "after_price":    after_price,
        "pct_change":     round(pct_change, 2),
        "days":           days,
    }


# ── MAIN ─────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("=== FCI Price Feed — fetching commodity prices ===\n")
    print("Fetching IMF prices...")
    imf = fetch_imf_prices(months_back=60)  # 5 years

    print("\nFetching World Bank Pink Sheet...")
    wb = fetch_wb_pink_sheet()

    # Merge: IMF primary, WB as supplement
    merged = {}
    for commodity in IMF_CODES:
        imf_data = imf.get(commodity, [])
        wb_data  = wb.get(commodity, [])
        # Prefer IMF, fill gaps with WB
        merged_dates = {p["date"]: p for p in wb_data}
        merged_dates.update({p["date"]: p for p in imf_data})
        merged[commodity] = sorted(merged_dates.values(), key=lambda x: x["date"])

    # Write price history
    history_path = OUT_DIR / "price-history.json"
    history_path.write_text(json.dumps({
        "generated_at": datetime.now().isoformat(),
        "sources":      ["IMF Primary Commodity Prices", "World Bank Pink Sheet"],
        "commodities":  merged,
    }, indent=2), encoding="utf-8")
    print(f"\nPrice history written: {history_path}")

    # Write latest prices
    latest = {}
    for commodity, prices in merged.items():
        if prices:
            latest[commodity] = prices[-1]
    latest_path = OUT_DIR / "price-latest.json"
    latest_path.write_text(json.dumps({
        "generated_at": datetime.now().isoformat(),
        "latest":       latest,
    }, indent=2), encoding="utf-8")
    print(f"Latest prices written: {latest_path}")

    print("\n=== SAMPLE: 30-day movement for March 2026 release ===")
    for commodity in IMF_CODES:
        result = get_price_movement(commodity, "2026-03-11", merged)
        print(f"  {commodity:12}: {result}")
