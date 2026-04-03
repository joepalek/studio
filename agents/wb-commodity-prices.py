"""
wb-commodity-prices.py — World Bank Pink Sheet Commodity Prices
Pool #58 | Tier 1 Pipeline Agent | Monthly cadence

Fetches the World Bank CMO Historical Data Monthly XLSX and extracts
agricultural commodity prices for Wheat, Corn, Soybeans, Cotton, Rice.

TASK SCHEDULER: \\Studio\\wb-commodity-prices | 03:00 daily | TTL 300s
"""

import json, os, sys, time, hashlib, io
from datetime import datetime
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.path.insert(0, "G:/My Drive/Projects/_studio")

AGENT_CONFIG = {
    "agent_id":      "wb-commodity-prices",
    "pool_number":   58,
    "pool_name":     "World Bank Pink Sheet Commodity Prices",
    "source_url":    "https://thedocs.worldbank.org/en/doc/18675f1d1639c7a34d463f59263ba0a2-0050012025/related/CMO-Historical-Data-Monthly.xlsx",
    "source_format": "xlsx",
    "cadence":       "monthly",
    "ttl_seconds":   300,
    "state_path":    "G:/My Drive/Projects/_studio/data/wb-commodity-prices/state.json",
    "digest_path":   "G:/My Drive/Projects/_studio/daily-digest.json",
    "log_path":      "G:/My Drive/Projects/_studio/logs/wb-commodity-prices.log",
    "commodities":   ["Wheat", "Corn", "Soybeans", "Cotton", "Rice"],
    # Pink sheet column keywords to match per commodity
    "synthesis_prompt": (
        "You are a commodity price analyst. Audience: food buyers tracking price trends.\n\n"
        "World Bank Pink Sheet price data:\n{data}\n\n"
        "For each commodity (Wheat, Corn, Soybeans, Cotton, Rice), write ONE sentence:\n"
        "  Format: 'CommodityName: [price trend description] -- [BULLISH/BEARISH/NEUTRAL] for buyers'\n\n"
        "BULLISH = prices rising (bad for buyers). BEARISH = prices falling (good for buyers).\n"
        "After the 5 lines, add: 'Overall: [BULLISH/BEARISH/NEUTRAL/MIXED] -- [one sentence]'\n"
        "Return ONLY these 6 lines. No preamble."
    ),
    "synthesize_max_tokens": 400,
    "commodity_map": {
        "Wheat":    ["wheat", "wht"],
        "Corn":     ["maize", "corn"],
        "Soybeans": ["soybean", "soybeans", "soy"],
        "Cotton":   ["cotton", "a-index"],
        "Rice":     ["rice"],
    },
}

# Column name keywords that indicate a price column (USD/unit)
_PRICE_KEYWORDS = ["price", "usd", "$/", "index", "cents"]
_DATE_KEYWORDS  = ["date", "period", "month", "year"]


def fetch_raw(cfg: dict) -> str:
    """Download XLSX, parse all sheets, return JSON string of raw rows."""
    import urllib.request, urllib.error, tempfile
    url = cfg["source_url"]
    max_failures = 3
    for attempt in range(1, max_failures + 1):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Studio/1.0"})
            with urllib.request.urlopen(req, timeout=30) as resp:
                data = resp.read()
            if len(data) < 500:
                raise ValueError(f"Response too small ({len(data)} bytes)")
            # Parse XLSX in memory
            try:
                import openpyxl
            except ImportError:
                raise RuntimeError("openpyxl not installed — run: pip install openpyxl")
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            all_rows = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                for row in rows:
                    if any(cell is not None for cell in row):
                        all_rows.append({
                            "sheet": sheet_name,
                            "values": [str(v) if v is not None else None for v in row],
                        })
            if not all_rows:
                raise ValueError("No rows extracted from XLSX")
            result = json.dumps(all_rows, ensure_ascii=False)
            return result
        except Exception as e:
            if attempt >= max_failures:
                raise
            time.sleep(5)
    raise RuntimeError("fetch_raw: unreachable")


def parse_raw(raw: str, cfg: dict) -> list:
    """
    Extract commodity price records from the raw row dump.
    Returns list of: {commodity, date, price_usd, unit, source}
    """
    rows = json.loads(raw)
    commodity_map = cfg.get("commodity_map", {})
    target_commodities = cfg.get("commodities", [])
    records = []

    # Find the "Monthly Prices" sheet or fall back to first sheet
    price_rows = [r for r in rows if "prices" in r["sheet"].lower() or "monthly" in r["sheet"].lower()]
    if not price_rows:
        price_rows = rows  # use all

    # Find the header row: pick the row (within first 20) with the most commodity keyword matches
    best_matches = 0
    header_idx = 0
    header = []
    for i, row in enumerate(price_rows[:20]):
        vals = [str(v).lower() if v else "" for v in row["values"]]
        matches = sum(
            1 for ckeys in commodity_map.values()
            for v in vals if any(ck in v for ck in ckeys)
        )
        if matches > best_matches:
            best_matches = matches
            header_idx = i
            header = vals

    # Map commodity → column index
    col_map = {}  # commodity_name → col_idx
    for commodity, keywords in commodity_map.items():
        if commodity not in target_commodities:
            continue
        for idx, h in enumerate(header):
            if any(kw in h for kw in keywords):
                col_map[commodity] = idx
                break

    # First column is typically the date
    date_col = 0
    for idx, h in enumerate(header):
        if any(dk in h for dk in _DATE_KEYWORDS):
            date_col = idx
            break

    # Extract data rows
    for row in price_rows[header_idx + 1:]:
        vals = row["values"]
        if not vals or all(v is None or v == "None" for v in vals):
            continue
        raw_date = vals[date_col] if date_col < len(vals) else None
        if not raw_date or raw_date == "None":
            continue
        for commodity, col_idx in col_map.items():
            if col_idx >= len(vals):
                continue
            raw_price = vals[col_idx]
            if raw_price is None or raw_price == "None":
                continue
            try:
                price = float(str(raw_price).replace(",", ""))
            except (ValueError, TypeError):
                continue
            records.append({
                "commodity":  commodity,
                "date":       str(raw_date)[:10],
                "price_usd":  price,
                "unit":       "USD/mt",
                "source":     "World Bank Pink Sheet",
            })

    # Deduplicate and sort by commodity + date
    seen = set()
    deduped = []
    for r in sorted(records, key=lambda x: (x["commodity"], x["date"])):
        key = (r["commodity"], r["date"])
        if key not in seen:
            seen.add(key)
            deduped.append(r)

    return deduped


def _load_state(cfg):
    try:
        return json.loads(Path(cfg["state_path"]).read_text(encoding="utf-8"))
    except Exception:
        return {"last_run": None, "last_checksum": None, "run_count": 0,
                "handoff_note": ""}


def _save_state(state, cfg):
    Path(cfg["state_path"]).parent.mkdir(parents=True, exist_ok=True)
    tmp = cfg["state_path"] + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(state, f, indent=2)
    os.replace(tmp, cfg["state_path"])


def run(cfg: dict, dry: bool = False) -> dict:
    import logging
    log = logging.getLogger(cfg["agent_id"])
    if not log.handlers:
        Path(cfg["log_path"]).parent.mkdir(parents=True, exist_ok=True)
        fh = logging.FileHandler(cfg["log_path"], encoding="utf-8")
        fh.setFormatter(logging.Formatter("[%(asctime)s] %(message)s", datefmt="%H:%M:%S"))
        log.addHandler(fh)
        ch = logging.StreamHandler(sys.stdout)
        ch.setFormatter(logging.Formatter("[%(asctime)s] %(message)s", datefmt="%H:%M:%S"))
        log.addHandler(ch)
        log.setLevel(logging.INFO)

    log.info(f"=== {cfg['agent_id']} START === pool #{cfg['pool_number']} | dry={dry}")
    state = _load_state(cfg)
    t0 = time.time()

    try:
        raw = fetch_raw(cfg)
        checksum = hashlib.md5(raw.encode()).hexdigest()[:12]

        if not dry and checksum == state.get("last_checksum"):
            log.info("HOPPER: no change since last run — skipping")
            state["last_run"] = datetime.now().isoformat()
            state["handoff_note"] = f"No change. checksum={checksum}"
            _save_state(state, cfg)
            return {"status": "skipped", "reason": "unchanged"}

        records = parse_raw(raw, cfg)
        log.info(f"Parsed {len(records)} price records across {len(set(r['commodity'] for r in records))} commodities")

        if not records:
            state["last_run"] = datetime.now().isoformat()
            state["handoff_note"] = "Empty parse"
            _save_state(state, cfg)
            return {"status": "empty", "records": 0}

        if not dry:
            state.update({
                "last_run":       datetime.now().isoformat(),
                "last_checksum":  checksum,
                "run_count":      state.get("run_count", 0) + 1,
                "record_count":   len(records),
                "handoff_note":   f"{len(records)} records | {len(set(r['commodity'] for r in records))} commodities | checksum={checksum}",
            })
            _save_state(state, cfg)
            log.info(f"state.json saved — run #{state['run_count']}")

        elapsed = round(time.time() - t0, 1)
        log.info(f"=== COMPLETE === {elapsed}s | {len(records)} records")
        return {"status": "ok", "records": len(records), "checksum": checksum}

    except Exception as e:
        log.error(f"PIPELINE FAILED: {e}")
        state["handoff_note"] = f"ERROR {datetime.now().isoformat()[:10]}: {str(e)[:60]}"
        _save_state(state, cfg)
        return {"status": "error", "error": str(e)}


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry", action="store_true")
    args = parser.parse_args()
    result = run(AGENT_CONFIG, dry=args.dry)
    print("Result:", result)
